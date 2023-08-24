#!/usr/bin/python
# -*- coding: utf-8 -*-
# @version        : 1.0
# @Create Time    : 2022/5/6 17:25 
# @File           : excel_manage.py
# @IDE            : PyCharm
# @desc           : EXCEL 文件操作

import datetime
import os
import re
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from application.settings import TEMP_DIR, TEMP_URL
import hashlib
import random
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from .excel_schema import AlignmentModel, FontModel, PatternFillModel


class ExcelManage:
    """
    excel 文件序列化
    """

    # 列名，A-Z
    EXCEL_COLUMNS = [chr(a) for a in range(ord('A'), ord('Z') + 1)]

    def __init__(self):
        self.sheet = None
        self.wb = None

    def open_workbook(self, file: str, read_only: bool = False, data_only: bool = False) -> None:
        """
        初始化 excel 文件

        :param file: 文件名称或者对象
        :param read_only: 是否只读，优化读取速度
        :param data_only: 是否加载文件对象
        """
        # 加载excel文件，获取表单
        self.wb = load_workbook(file, read_only=read_only, data_only=data_only)

    def open_sheet(self, sheet_name: str = None, **kwargs) -> None:
        """
        初始化 excel 文件

        :param sheet_name: 表单名称，为空则默认第一个
        """
        # 加载excel文件，获取表单
        if not self.wb:
            self.open_workbook(kwargs.get("file"), kwargs.get("read_only", False), kwargs.get("data_only", False))
        if sheet_name:
            self.sheet = self.wb[sheet_name]
        else:
            self.sheet = self.wb.active

    def get_sheets(self) -> list:
        """
        读取所有工作区名称
        :return: 一维数组
        """
        return self.wb.sheetnames

    def create_excel(self, sheet_name: str = None) -> None:
        """
        创建 excel 文件

        :param sheet_name: 表单名称，为空则默认第一个
        """
        # 加载excel文件，获取表单
        self.wb = Workbook()
        self.sheet = self.wb.active
        if sheet_name:
            self.sheet.title = sheet_name

    def readlines(self, min_row: int = 1, min_col: int = 1, max_row: int = None, max_col: int = None) -> list:
        """
        读取指定表单所有数据

        :return: 二维数组
        """
        rows = self.sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)
        result = []
        for row in rows:
            _row = []
            for cell in row:
                _row.append(cell.value)
            if any(_row):
                result.append(_row)
        return result

    def get_header(self, row: int = 1, col: int = None, asterisk: bool = False) -> list:
        """
        读取指定表单的表头（第一行数据）

        :param row: 指定行
        :param col: 最大列
        :param asterisk: 是否去除 * 号
        :return: 一维数组
        """
        rows = self.sheet.iter_rows(min_row=row, max_col=col)
        result = []
        for row in rows:
            for cell in row:
                value = cell.value.replace("*", "").strip() if asterisk else cell.value.strip()
                result.append(value)
            break
        return result

    def write_list(self, rows: list, header: list = None) -> None:
        """
        写入 excel文件

        :param rows: 行数据集
        :param header: 表头
        """
        if header:
            self.sheet.append(header)
            pattern_fill_style = PatternFillModel(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            font_style = FontModel(bold=True)
            self.__set_row_style(1, len(header), pattern_fill_style=pattern_fill_style, font_style=font_style)
        for index, data in enumerate(rows):
            format_columns = {
                "date_columns": []
            }
            for i in range(0, len(data)):
                if isinstance(data[i], datetime.datetime):
                    data[i] = data[i].strftime("%Y/%m/%d %H:%M:%S")
                    format_columns["date_columns"].append(i + 1)
            self.sheet.append(data)
            self.__set_row_style(index + 2, len(data))
            self.__set_row_format(index + 2, format_columns)
        self.__auto_width()
        self.__set_row_border()

    def save_excel(self, filename: str = None):
        """
        保存 excel 文件到本地
        默认保存到临时目录中
        """
        date = datetime.datetime.strftime(datetime.datetime.now(), "%Y%m%d")
        file_dir = os.path.join(TEMP_DIR, date)
        if not os.path.exists(file_dir):
            os.mkdir(file_dir)
        if not filename:
            name = hashlib.md5(str(random.random()).encode()).hexdigest() + ".xlsx"
            filename = os.path.join(file_dir, name)
        else:
            name = filename
            filename = os.path.join(file_dir, filename)
        self.wb.save(filename)
        # 返回访问路由
        return f"{TEMP_URL}/{date}/{name}"

    def __set_row_style(
            self,
            row: int,
            max_column: int,
            alignment_style: AlignmentModel = AlignmentModel(),
            font_style: FontModel = FontModel(),
            pattern_fill_style: PatternFillModel = PatternFillModel()
    ):
        """
        设置行样式

        :param row: 行
        :param max_column: 最大列
        :param alignment_style: 单元格内容的对齐设置
        :param font_style: 单元格内容的字体样式设置
        :param pattern_fill_style: 单元格的填充模式设置
        """
        for index in range(0, max_column):
            alignment = Alignment(**alignment_style.model_dump())
            font = Font(**font_style.model_dump())
            pattern_fill = PatternFill(**pattern_fill_style.model_dump())
            self.sheet.cell(row=row, column=index + 1).alignment = alignment
            self.sheet.cell(row=row, column=index + 1).font = font
            self.sheet.cell(row=row, column=index + 1).fill = pattern_fill

    def __set_row_format(self, row: int, columns: dict):
        """
        格式化行数据类型

        :param row: 行
        :param columns: 列数据
        """
        for index in columns.get("date_columns", []):
            self.sheet.cell(row=row, column=index).number_format = "yyyy/mm/dd h:mm:ss"

    def __set_row_border(self):
        """
        设置行边框
        """
        # 创建 Border 对象并设置边框样式
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        # 设置整个表格的边框
        for row in self.sheet.iter_rows():
            for cell in row:
                cell.border = border

    def __auto_width(self):
        """
        设置自适应列宽
        """
        # 设置一个字典用于保存列宽数据
        dims = {}
        # 遍历表格数据，获取自适应列宽数据
        for row in self.sheet.rows:
            for cell in row:
                if cell.value:
                    # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                    # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                    # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                    cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                    dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
        for col, value in dims.items():
            # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
            self.sheet.column_dimensions[get_column_letter(col)].width = value + 10

    def close(self):
        """
        关闭文件
        """
        self.wb.close()


if __name__ == '__main__':
    print([chr(a) for a in range(ord('A'), ord('Z') + 1)])
